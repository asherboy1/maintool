# coding: utf-8
import os
import xlrd
import xlwt
from datetime import datetime
from xlutils.copy import copy as copyxls
from openpyxl import load_workbook, Workbook
from win32com.client import DispatchEx

class ExcelHandler(object):
    def __init__(self, filename, refresh=False) -> None:
        if os.path.exists(filename):
            if refresh:
                self.refresh_all(filename)
            self.workbook = xlrd.open_workbook(filename)
            self.sheets = self.workbook.sheet_names()

            self.current_sheet = self.workbook.sheet_by_index(0)
            self.filename = filename
        else:
            raise ValueError("file missing!")
    
    @staticmethod
    def refresh_all(filename):
        excel = DispatchEx('Excel.Application')
        excel.Visible = False
        workbook = excel.WorkBooks.Open(filename)
        workbook.RefreshAll()
        workbook.Save()
        workbook.Close(True)
        excel.Quit()
    
    def read_all(self, sheet_name=None):
        if sheet_name is not None:
            self.current_sheet = self.workbook.sheet_by_name(sheet_name)
        results = [self.current_sheet.row(i) for i in range(self.current_sheet.nrows)]
        for row, row_range, col, col_range in self.current_sheet.merged_cells:
            for x in range(row, row_range):
                for y in range(col, col_range):
                    results[x][y] = self.current_sheet.cell(row, col)
        values = []
        for row in range(self.current_sheet.nrows):
            values.append([self._format_value(cell) for cell in results[row]])
        return values

    def _format_value(self, cell):
        if cell.ctype == 2 and cell.value % 1 == 0:  # 如果为整形
            return int(cell.value)
        elif cell.ctype == 3:
            # 转成datetime对象
            try:
                date = datetime(*xlrd.xldate_as_tuple(cell.value, self.workbook.datemode))
            except ValueError:
                date_tuple = xlrd.xldate_as_tuple(cell.value, self.workbook.datemode)
                default_date = [1900, 1, 1]
                default_date.extend(date_tuple[3:])
                date = datetime(*tuple(default_date))
                return date.strftime("%H:%M:%S")
            return date.strftime("%Y-%m-%d %H:%M:%S")
        elif cell.ctype == 4:
            return cell.value == 1
        else:
            return cell.value
    
    def read_row(self, sheet_name=None, row=0, row_range=1):
        self.current_sheet = self.sheet(sheet_name)
        results = self.read_all(sheet_name)
        return results[row:row + row_range]
    
    def read_row_until_blank(self, sheet_name=None, row=0):
        self.current_sheet = self.sheet(sheet_name)
        results = self.read_all(sheet_name)[row:]
        index = 0
        for i, row in enumerate(results):
            if not len([v for v in filter(lambda r: r, row)]):
                index = i 
                break
        return results[:index]
    
    def read_dict_row_until_blank(self, sheet_name=None, row=0):
        self.current_sheet = self.sheet(name=sheet_name)
        results = self.read_all(sheet_name)[row:]
        datas = []
        for i, row in enumerate(results):
            if not len([v for v in filter(lambda r:r, row)]):
                break
            elif i == 0:
                continue
            datas.append({k: v for k, v in zip(results[0], row)})
        return datas

    def read_colum_by_cell_value(self, value, sheet_name=None):
        self.current_sheet = self.sheet(name=sheet_name)
        x, y = self.find_cell_by_value(value)
        values = map(self._format_value,
                         [self.current_sheet.cell(row, y) for row in range(x + 1, self.current_sheet.nrows)])
        return [v for v in filter(lambda v: v, values)]

    def find_cell_by_value(self, value):
        for i in range(self.current_sheet.nrows):
            values = self.current_sheet.row_values(i, 0, self.current_sheet.ncols)
            if value in values:
                return i, values.index(value)
        return -1, -1
    
    def sheet(self, index=None, name=None):
        if index is None and name:
            return self.workbook.sheet_by_name(name)
        elif index is not None and name is None:
            return self.workbook.sheet_by_index(index)
        elif index is None and name is None:
            return self.workbook.sheet_by_index(0)
        else:
            raise IndexError("索引或名称错误")
    
    def read_cell(self, sheet_name, row, col):
        self.current_sheet = self.workbook.sheet_by_name(sheet_name)
        if row >= self.current_sheet.nrows or col >= self.current_sheet.ncols:
            raise IndexError("索引超出范围！")
        else:
            cell = self.current_sheet.cell(row, col)
            return self._format_value(cell) if cell.ctype != 0 else self._format_value(
                self._try_merged_cell(row, col, cell)
            )
    
    def _try_merged_cell(self, x, y, cell):
        for row, row_range, col, col_range in self.current_sheet.merged_cells:
            if row <= x < row_range and col <= y <= col_range:
                cell = self.current_sheet.cell(row, col)
            else:
                continue
        return cell
    
    def replace_xls(self):
        self.refresh_all(self.filename)
        wb = load_workbook(self.filename, data_only=True)
        name, ext = os.path.splitext(self.filename)
        new_file_name = name + "_copy" + ext
        f = open(new_file_name, mode="w", encoding='GBK')
        f.close()
        wb2 = Workbook()
        for sheetname in self.workbook.sheet_names():
            ws = wb[sheetname]
            ws2 = wb2.create_sheet(title=sheetname)
            for i, row in enumerate(ws.iter_rows()):
                for j, cell in enumerate(row):
                    ws2.cell(row=i + 1, column=j + 1, value=getattr(cell, 'value', None))
        wb2.remove(wb2['Sheet'])
        wb2.save(new_file_name)
        return new_file_name

    def write_data_with_key(self, sheetname='Sheet1', datas=None, copy=True):
        self.current_sheet = self.sheet(name=sheetname)
        wb = load_workbook(self.filename)
        ws = wb.get_sheet_by_name(sheetname)
        for i, k in enumerate(datas[0].keys()):
            ws.cell(row=1, colnum=i+1, value=k)
        for index, data in enumerate(datas):
            for k, v in data.items():
                r, c = self.find_cell_by_value(k)
                ws.cell(row=r + index + 2, column=c + 1, value=v)
        if copy:
            name, ext = os.path.splitext(self.filename)
            new_file_name = name + "_copy" + ext
        else:
            new_file_name = self.filename
        wb.save(new_file_name)
        return new_file_name
    
    def rename_sheet(self, name, new_name):
        wb = load_workbook(self.filename, data_only=True)
        wb.get_sheet_by_name(name).title = new_name
        wb.save(self.filename)

    def delete_from_excel(self, sheet, start_row=0, end_row=None):
        wb = load_workbook(self.filename, data_only=True)
        ws = wb.get_sheet_by_name(name=sheet)
        end_row = ws.max_row if end_row is None else end_row
        ws.delete_rows(start_row+1, end_row-start_row+1)
        wb.save(self.filename)

def write_excel_xls(path, sheet_name, value):
    index = len(value) # 获取行数
    Workbook = xlwt.Workbook()
    sheet = Workbook.add_sheet(sheet_name) # 新建表格
    for i in range(0, index):
        for j in range(0, len(value(i))):
            sheet.write(i, j, value[i][j])
    Workbook.save(path)
    return print("ok")

def write_excel_xls_append(path, value):
    index = len(value)
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    rows_old = worksheet.nrows
    new_workbook = copyxls(workbook)
    new_worksheet = new_workbook.get_sheet(0)

    for i in range(0, index):
        for j in range(0, len(value[i])):
            new_worksheet.write(o + rows_old, j, value[i][j])
    
    new_workbook.save(path)
    return print('ok')






















